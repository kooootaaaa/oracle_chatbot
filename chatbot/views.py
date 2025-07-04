from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
import json
import html
from .models import TFurenSho
from .forms import ExportForm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re

def index(request):
    return render(request, 'chatbot/index.html')

@csrf_exempt
def chat(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_input = data.get('message', '').strip()
            action = data.get('action', 'search')  # ãƒ‡ãƒ•ã‚©ãƒ«ãƒˆã¯æ¤œç´¢
            page = data.get('page', 1)  # ãƒšãƒ¼ã‚¸ç•ªå·ï¼ˆ1ã‹ã‚‰é–‹å§‹ï¼‰
            
            if not user_input:
                return JsonResponse({'response': 'è³ªå•ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚'})
            
            # æ¤œç´¢ã‚¯ã‚¨ãƒªã®æ§‹ç¯‰ï¼ˆANDæ¤œç´¢ï¼šã™ã¹ã¦ã®ã‚­ãƒ¼ãƒ¯ãƒ¼ãƒ‰ã‚’å«ã‚€ãƒ¬ã‚³ãƒ¼ãƒ‰ï¼‰
            import re
            # èª­ç‚¹ã€å…¨è§’ã‚«ãƒ³ãƒã€åŠè§’ã‚«ãƒ³ãƒã€ã‚¹ãƒšãƒ¼ã‚¹ã§åˆ†å‰²ã—ã¦ã‚­ãƒ¼ãƒ¯ãƒ¼ãƒ‰ã‚’æŠ½å‡º
            keywords = re.split(r'[,ã€ï¼Œ\s]+', user_input.strip())
            keywords = [k.strip() for k in keywords if k.strip()]  # ç©ºæ–‡å­—åˆ—ã‚’é™¤å»
            
            query = Q()
            for keyword in keywords:
                # å„ã‚­ãƒ¼ãƒ¯ãƒ¼ãƒ‰ã«ã¤ã„ã¦ã€ã„ãšã‚Œã‹ã®ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã«å«ã¾ã‚Œã‚‹æ¡ä»¶
                keyword_query = (
                    Q(hno__icontains=keyword) |
                    Q(keikaku_no__icontains=keyword) |
                    Q(tr_hinban__icontains=keyword) |
                    Q(tr_hinmei__icontains=keyword) |
                    Q(shashu__icontains=keyword) |
                    Q(fuguai_naiyou__icontains=keyword) |
                    Q(suitei_fuguai_genin__icontains=keyword) |
                    Q(taisaku_an__icontains=keyword)
                )
                # ã™ã¹ã¦ã®ã‚­ãƒ¼ãƒ¯ãƒ¼ãƒ‰æ¡ä»¶ã‚’ANDã§çµåˆ
                if query:
                    query &= keyword_query
                else:
                    query = keyword_query
            
            print(f"Original input: {user_input}")  # ãƒ‡ãƒãƒƒã‚°ç”¨
            print(f"Keywords: {keywords}")  # ãƒ‡ãƒãƒƒã‚°ç”¨
            print(f"Query: {query}")  # ãƒ‡ãƒãƒƒã‚°ç”¨
            
            # ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹æ¤œç´¢
            if action == 'detail':
                # è©³ç´°è¡¨ç¤ºã®å ´åˆï¼ˆè¨ˆç”»æ›¸NOé™é †ï¼‰
                detail_type = data.get('detail_type', '')
                results = TFurenSho.objects.filter(query).order_by('-hno')[:10]
                
                if results:
                    response_text = '<div class="detail-content">'
                    for i, result in enumerate(results, 1):
                        response_text += f'<div class="detail-section">'
                        response_text += f'<h3>æ¤œç´¢çµæœ {i}</h3>'
                        response_text += f'<div class="detail-label">è©•æŠ€é€£No.:</div><div>{result.hno or "---"}</div><br>'
                        response_text += f'<div class="detail-label">è¨ˆç”»æ›¸No.:</div><div>{result.keikaku_no or "---"}</div><br>'
                        response_text += f'<div class="detail-label">TRå“ç•ª:</div><div>{result.tr_hinban or "---"}</div><br>'
                        response_text += f'<div class="detail-label">TRå“å:</div><div>{result.tr_hinmei or "---"}</div><br>'
                        response_text += f'<div class="detail-label">è»Šç¨®:</div><div>{result.shashu or "---"}</div><br>'
                        response_text += f'<div class="detail-label">ä¸å…·åˆå†…å®¹:</div><div style="white-space: pre-wrap;">{result.fuguai_naiyou or "---"}</div><br>'
                        response_text += f'<div class="detail-label">æ¨å®šä¸å…·åˆåŸå› :</div><div style="white-space: pre-wrap;">{result.suitei_fuguai_genin or "---"}</div><br>'
                        response_text += f'<div class="detail-label">å¯¾ç­–æ¡ˆ:</div><div style="white-space: pre-wrap;">{result.taisaku_an or "---"}</div>'
                        response_text += '</div>'
                    response_text += '</div>'
                else:
                    response_text = '<p>è©²å½“ã™ã‚‹è©³ç´°æƒ…å ±ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚</p>'
                
                return JsonResponse({'response': response_text})
            
            elif action == 'summary':
                # è¦ç´„ã®å ´åˆã¯ã™ã¹ã¦å–å¾—ï¼ˆè¨ˆç”»æ›¸NOé™é †ï¼‰
                results = TFurenSho.objects.filter(query).order_by('-hno')
                
                if results:
                    # ãƒ‡ãƒ¼ã‚¿åˆ†æ
                    hno_list = [r.hno for r in results if r.hno]
                    keikaku_no_list = [r.keikaku_no for r in results if r.keikaku_no]
                    tr_hinban_set = set([r.tr_hinban for r in results if r.tr_hinban])
                    tr_hinmei_set = set([r.tr_hinmei for r in results if r.tr_hinmei])
                    shashu_set = set([r.shashu for r in results if r.shashu])
                    
                    # ä¸å…·åˆå†…å®¹ã€åŸå› ã€å¯¾ç­–ã®è¦ç´„ï¼ˆæœ€åˆã®3ä»¶ã‹ã‚‰æŠœç²‹ï¼‰
                    fuguai_samples = [r.fuguai_naiyou for r in results[:3] if r.fuguai_naiyou]
                    genin_samples = [r.suitei_fuguai_genin for r in results[:3] if r.suitei_fuguai_genin]
                    taisaku_samples = [r.taisaku_an for r in results[:3] if r.taisaku_an]
                    
                    # è¦ç´„ãƒ¬ã‚¹ãƒãƒ³ã‚¹ä½œæˆ
                    response_text = f"ã€Œ{user_input}ã€å…¨{len(results)}ä»¶ã®è¦ç´„\n\n"
                    
                    # è©•æŠ€é€£No.
                    if len(hno_list) > 5:
                        response_text += f"è©•æŠ€é€£No.: {hno_list[0]}ï½{hno_list[-1]} ä»–{len(hno_list)-2}ä»¶\n"
                    else:
                        response_text += f"è©•æŠ€é€£No.: {', '.join(hno_list)}\n"
                    
                    # è¨ˆç”»æ›¸No.
                    if keikaku_no_list:
                        if len(keikaku_no_list) > 5:
                            response_text += f"è¨ˆç”»æ›¸No.: {keikaku_no_list[0]}ï½{keikaku_no_list[-1]} ä»–{len(keikaku_no_list)-2}ä»¶\n"
                        else:
                            response_text += f"è¨ˆç”»æ›¸No.: {', '.join(keikaku_no_list)}\n"
                    else:
                        response_text += "è¨ˆç”»æ›¸No.: ãƒ‡ãƒ¼ã‚¿ãªã—\n"
                    
                    # TRå“ç•ª
                    if len(tr_hinban_set) > 3:
                        tr_list = sorted(tr_hinban_set)
                        response_text += f"TRå“ç•ª: {', '.join(tr_list[:3])} ä»–{len(tr_hinban_set)-3}ç¨®é¡\n"
                    else:
                        response_text += f"TRå“ç•ª: {', '.join(sorted(tr_hinban_set))}\n"
                    
                    # TRå“å
                    if len(tr_hinmei_set) > 3:
                        hinmei_list = sorted(tr_hinmei_set)
                        response_text += f"TRå“å: {', '.join(hinmei_list[:3])} ä»–{len(tr_hinmei_set)-3}ç¨®é¡\n"
                    else:
                        response_text += f"TRå“å: {', '.join(sorted(tr_hinmei_set))}\n"
                    
                    # è»Šç¨®
                    if len(shashu_set) > 3:
                        shashu_list = sorted(shashu_set)
                        response_text += f"è»Šç¨®: {', '.join(shashu_list[:3])} ä»–{len(shashu_set)-3}ç¨®é¡\n\n"
                    else:
                        response_text += f"è»Šç¨®: {', '.join(sorted(shashu_set))}\n\n"
                    
                    # ä¸å…·åˆå†…å®¹è¦ç´„
                    response_text += "ã€ä¸»ãªä¸å…·åˆå†…å®¹ã€‘\n"
                    for i, fuguai in enumerate(fuguai_samples, 1):
                        response_text += f"{i}. {fuguai[:100]}{'...' if len(fuguai) > 100 else ''}\n"
                    
                    response_text += "\nã€ä¸»ãªæ¨å®šåŸå› ã€‘\n"
                    for i, genin in enumerate(genin_samples, 1):
                        response_text += f"{i}. {genin[:100]}{'...' if len(genin) > 100 else ''}\n"
                    
                    response_text += "\nã€ä¸»ãªå¯¾ç­–æ¡ˆã€‘\n"
                    for i, taisaku in enumerate(taisaku_samples, 1):
                        response_text += f"{i}. {taisaku[:100]}{'...' if len(taisaku) > 100 else ''}\n"
                    
                    return JsonResponse({'response': response_text})
                else:
                    return JsonResponse({'response': 'è¦ç´„å¯¾è±¡ã®ãƒ‡ãƒ¼ã‚¿ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚'})
            else:
                # é€šå¸¸æ¤œç´¢ã¯ãƒšãƒ¼ã‚¸ãƒãƒ¼ã‚·ãƒ§ãƒ³å¯¾å¿œï¼ˆ10ä»¶ãšã¤ï¼‰è¨ˆç”»æ›¸NOé™é †
                total_count = TFurenSho.objects.filter(query).count()
                start_index = (page - 1) * 10
                end_index = start_index + 10
                results = TFurenSho.objects.filter(query).order_by('-hno')[start_index:end_index]
            
            if results:
                if action == 'summary':
                    # è¦ç´„ãƒ¢ãƒ¼ãƒ‰ã®å‡¦ç†
                    response_text = f"ã€Œ{user_input}ã€ã«é–¢é€£ã™ã‚‹ä¸å…·åˆæƒ…å ±ã®è¦ç´„:\n\n"
                    response_text += f"ã€æ¤œç´¢çµæœã€‘ {len(results)}ä»¶\n\n"
                    
                    # ä¸å…·åˆå†…å®¹ã®å‚¾å‘ã‚’ã¾ã¨ã‚ã‚‹
                    fuguai_list = []
                    genin_list = []
                    taisaku_list = []
                    shashu_set = set()
                    
                    for result in results:
                        if result.fuguai_naiyou:
                            fuguai_list.append(result.fuguai_naiyou[:100])  # æœ€åˆã®100æ–‡å­—
                        if result.suitei_fuguai_genin:
                            genin_list.append(result.suitei_fuguai_genin[:100])
                        if result.taisaku_an:
                            taisaku_list.append(result.taisaku_an[:100])
                        if result.shashu:
                            shashu_set.add(result.shashu)
                    
                    # è¦ç´„ç”¨ã®ãƒ‡ãƒ¼ã‚¿ã‚’æ ¼ç´
                    summary_data = {
                        'keyword': user_input,
                        'results': results
                    }
                    
                    response_text += f"ã€é–¢é€£è»Šç¨®ã€‘\n"
                    if shashu_set:
                        for shashu in sorted(shashu_set):
                            # HTMLã‚¨ã‚¹ã‚±ãƒ¼ãƒ—ã‚’å‰Šé™¤ã—ã¦ã‚¯ãƒªãƒƒã‚¯å¯èƒ½ã«ã™ã‚‹
                            safe_shashu = shashu.replace("'", "\\'").replace('"', '\\"')
                            response_text += f'<span class="clickable-item" onclick="showDetail(\'{safe_shashu}\', \'shashu\')">{shashu}</span>ã€€'
                        response_text += "\n\n"
                    else:
                        response_text += "è»Šç¨®æƒ…å ±ãªã—\n\n"
                    
                    response_text += f"ã€ä¸»ãªä¸å…·åˆå†…å®¹ã€‘ï¼ˆã‚¯ãƒªãƒƒã‚¯ã§è©³ç´°è¡¨ç¤ºï¼‰\n"
                    for i, fuguai in enumerate(fuguai_list[:3], 1):
                        # HTMLã‚¨ã‚¹ã‚±ãƒ¼ãƒ—ã‚’å‰Šé™¤ã—ã¦ã‚¯ãƒªãƒƒã‚¯å¯èƒ½ã«ã™ã‚‹
                        safe_fuguai = fuguai.replace("'", "\\'").replace('"', '\\"').replace('\n', ' ')
                        response_text += f'{i}. <span class="clickable-item" onclick="showDetailFromSummary({i-1}, \'fuguai\')">{fuguai}</span>...\n'
                    
                    response_text += f"\nã€ä¸»ãªæ¨å®šåŸå› ã€‘ï¼ˆã‚¯ãƒªãƒƒã‚¯ã§è©³ç´°è¡¨ç¤ºï¼‰\n"
                    for i, genin in enumerate(genin_list[:3], 1):
                        # HTMLã‚¨ã‚¹ã‚±ãƒ¼ãƒ—ã‚’å‰Šé™¤ã—ã¦ã‚¯ãƒªãƒƒã‚¯å¯èƒ½ã«ã™ã‚‹
                        safe_genin = genin.replace("'", "\\'").replace('"', '\\"').replace('\n', ' ')
                        response_text += f'{i}. <span class="clickable-item" onclick="showDetailFromSummary({i-1}, \'genin\')">{genin}</span>...\n'
                    
                    response_text += f"\nã€ä¸»ãªå¯¾ç­–æ¡ˆã€‘ï¼ˆã‚¯ãƒªãƒƒã‚¯ã§è©³ç´°è¡¨ç¤ºï¼‰\n"
                    for i, taisaku in enumerate(taisaku_list[:3], 1):
                        # HTMLã‚¨ã‚¹ã‚±ãƒ¼ãƒ—ã‚’å‰Šé™¤ã—ã¦ã‚¯ãƒªãƒƒã‚¯å¯èƒ½ã«ã™ã‚‹
                        safe_taisaku = taisaku.replace("'", "\\'").replace('"', '\\"').replace('\n', ' ')
                        response_text += f'{i}. <span class="clickable-item" onclick="showDetailFromSummary({i-1}, \'taisaku\')">{taisaku}</span>...\n'
                    
                    # è¦ç´„ãƒ‡ãƒ¼ã‚¿ã‚’ä¿å­˜ï¼ˆã‚»ãƒƒã‚·ãƒ§ãƒ³ã¾ãŸã¯ã‚­ãƒ£ãƒƒã‚·ãƒ¥ã«ä¿å­˜ã™ã‚‹ã“ã¨ã‚‚å¯èƒ½ï¼‰
                    request.session['summary_results'] = [
                        {
                            'hno': r.hno,
                            'tr_hinban': r.tr_hinban,
                            'tr_hinmei': r.tr_hinmei,
                            'shashu': r.shashu,
                            'fuguai_naiyou': r.fuguai_naiyou,
                            'suitei_fuguai_genin': r.suitei_fuguai_genin,
                            'taisaku_an': r.taisaku_an
                        } for r in results
                    ]
                    
                else:
                    # é€šå¸¸æ¤œç´¢ãƒ¢ãƒ¼ãƒ‰ã®å‡¦ç†ï¼ˆãƒšãƒ¼ã‚¸ãƒãƒ¼ã‚·ãƒ§ãƒ³å¯¾å¿œï¼‰
                    current_start = (page - 1) * 10 + 1
                    current_end = min(page * 10, total_count)
                    total_pages = (total_count + 9) // 10  # åˆ‡ã‚Šä¸Šã’è¨ˆç®—
                    
                    response_text = f"ã€Œ{user_input}ã€ã®æ¤œç´¢çµæœ: {current_start}-{current_end}ä»¶ / å…¨{total_count}ä»¶\n\n"
                    
                    for i, result in enumerate(results, current_start):
                        response_text += f"ã€{i}ã€‘\n"
                        response_text += f"è©•æŠ€é€£No.: {result.hno}\n"
                        response_text += f"è¨ˆç”»æ›¸No.: {result.keikaku_no or '---'}\n"
                        response_text += f"TRå“ç•ª: {result.tr_hinban}\n"
                        response_text += f"TRå“å: {result.tr_hinmei}\n"
                        response_text += f"è»Šç¨®: {result.shashu}\n"
                        response_text += f"ä¸å…·åˆå†…å®¹: {result.fuguai_naiyou}\n"
                        response_text += f"æ¨å®šä¸å…·åˆåŸå› : {result.suitei_fuguai_genin}\n"
                        response_text += f"å¯¾ç­–æ¡ˆ: {result.taisaku_an}\n"
                        response_text += "-" * 50 + "\n"
                    
                    # ãƒšãƒ¼ã‚¸ãƒãƒ¼ã‚·ãƒ§ãƒ³ã¨è¦ç´„ãƒœã‚¿ãƒ³ã‚’è¿½åŠ 
                    response_text += '\n<div class="pagination-container">'
                    
                    if page > 1:
                        response_text += f'<button class="btn pagination-btn" onclick="searchPage(\'{user_input}\', {page - 1})">â† å‰ã®10ä»¶</button>'
                    if page < total_pages:
                        response_text += f'<button class="btn pagination-btn" onclick="searchPage(\'{user_input}\', {page + 1})">æ¬¡ã®10ä»¶ â†’</button>'
                    
                    # è¦ç´„ãƒœã‚¿ãƒ³ã‚’è¿½åŠ 
                    safe_keyword = user_input.replace("'", "\\'").replace('"', '\\"')
                    response_text += f'<button class="btn summary-btn" onclick="showSummary(\'{safe_keyword}\')">ğŸ“Š è¦ç´„</button>'
                    
                    response_text += '</div>'
            else:
                response_text = f"ç”³ã—è¨³ã”ã–ã„ã¾ã›ã‚“ãŒã€ã€Œ{user_input}ã€ã«é–¢é€£ã™ã‚‹æƒ…å ±ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚\n\nTRå“ç•ªã€TRå“åã€è»Šç¨®åãªã©ã§æ¤œç´¢ã—ã¦ã¿ã¦ãã ã•ã„ã€‚"
            
            return JsonResponse({'response': response_text})
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in chat view: {error_details}")  # ã‚µãƒ¼ãƒãƒ¼ãƒ­ã‚°ã«å‡ºåŠ›
            return JsonResponse({'response': f'ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {str(e)}'})
    
    return JsonResponse({'response': 'POSTãƒªã‚¯ã‚¨ã‚¹ãƒˆã®ã¿å¯¾å¿œã—ã¦ã„ã¾ã™ã€‚'})

def export_view(request):
    """Excelå‡ºåŠ›ç”¨ãƒ“ãƒ¥ãƒ¼"""
    if request.method == 'POST':
        form = ExportForm(request.POST)
        if form.is_valid():
            # ãƒ•ã‚©ãƒ¼ãƒ ã‹ã‚‰ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
            selected_fields = form.cleaned_data['fields']
            search_query = form.cleaned_data['search_query']
            max_records = form.cleaned_data['max_records']
            
            # ã‚¯ã‚¨ãƒªã®æ§‹ç¯‰
            query = Q()
            if search_query:
                keywords = re.split(r'[,ã€ï¼Œ\s]+', search_query.strip())
                keywords = [k.strip() for k in keywords if k.strip()]
                
                for keyword in keywords:
                    query &= (
                        Q(hno__icontains=keyword) |
                        Q(keikaku_no__icontains=keyword) |
                        Q(tr_hinban__icontains=keyword) |
                        Q(tr_hinmei__icontains=keyword) |
                        Q(shashu__icontains=keyword) |
                        Q(fuguai_naiyou__icontains=keyword) |
                        Q(suitei_fuguai_genin__icontains=keyword) |
                        Q(taisaku_an__icontains=keyword)
                    )
            
            # ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
            results = TFurenSho.objects.filter(query)[:max_records]
            
            # Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ä½œæˆ
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'T_ä¸é€£æ›¸ãƒ‡ãƒ¼ã‚¿'
            
            # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã®ã‚¹ã‚¿ã‚¤ãƒ«è¨­å®š
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰åã¨ãƒ˜ãƒƒãƒ€ãƒ¼ãƒ†ã‚­ã‚¹ãƒˆã®ãƒãƒƒãƒ”ãƒ³ã‚°
            field_headers = {
                'hno': 'è©•æŠ€é€£No. (HNO)',
                'keikaku_no': 'è¨ˆç”»æ›¸NO',
                'tr_hinban': 'TRå“ç•ª',
                'tr_hinmei': 'TRå“å',
                'shashu': 'è»Šç¨®',
                'fuguai_naiyou': 'ä¸å…·åˆå†…å®¹',
                'suitei_fuguai_genin': 'æ¨å®šä¸å…·åˆåŸå› ',
                'taisaku_an': 'å¯¾ç­–æ¡ˆç­‰',
            }
            
            # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã‚’ä½œæˆ
            headers = [field_headers[field] for field in selected_fields]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # ãƒ‡ãƒ¼ã‚¿è¡Œã‚’è¿½åŠ 
            for row_num, result in enumerate(results, 2):
                for col_num, field in enumerate(selected_fields, 1):
                    value = getattr(result, field, '')
                    ws.cell(row=row_num, column=col_num, value=value or '')
            
            # åˆ—å¹…ã®è‡ªå‹•èª¿æ•´
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # HTTPãƒ¬ã‚¹ãƒãƒ³ã‚¹ã‚’ä½œæˆ
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'T_ä¸é€£æ›¸_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            # Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ä¿å­˜
            wb.save(response)
            return response
    else:
        form = ExportForm()
    
    return render(request, 'chatbot/export.html', {'form': form})

@csrf_exempt
def detail(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            index = data.get('index', 0)
            detail_type = data.get('type', '')
            
            # ã‚»ãƒƒã‚·ãƒ§ãƒ³ã‹ã‚‰è¦ç´„çµæœã‚’å–å¾—
            summary_results = request.session.get('summary_results', [])
            
            if 0 <= index < len(summary_results):
                result = summary_results[index]
                
                response_text = '<div class="detail-content">'
                response_text += f'<div class="detail-section">'
                response_text += f'<div class="detail-label">TRå“ç•ª:</div><div>{html.escape(result.get("tr_hinban", "") or "---")}</div><br>'
                response_text += f'<div class="detail-label">TRå“å:</div><div>{html.escape(result.get("tr_hinmei", "") or "---")}</div><br>'
                response_text += f'<div class="detail-label">è»Šç¨®:</div><div>{html.escape(result.get("shashu", "") or "---")}</div><br>'
                response_text += f'<div class="detail-label">ä¸å…·åˆå†…å®¹:</div><div style="white-space: pre-wrap;">{html.escape(result.get("fuguai_naiyou", "") or "---")}</div><br>'
                response_text += f'<div class="detail-label">æ¨å®šä¸å…·åˆåŸå› :</div><div style="white-space: pre-wrap;">{html.escape(result.get("suitei_fuguai_genin", "") or "---")}</div><br>'
                response_text += f'<div class="detail-label">å¯¾ç­–æ¡ˆ:</div><div style="white-space: pre-wrap;">{html.escape(result.get("taisaku_an", "") or "---")}</div>'
                response_text += '</div>'
                response_text += '</div>'
            else:
                response_text = '<p>è©³ç´°æƒ…å ±ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚</p>'
                
            return JsonResponse({'response': response_text})
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in detail view: {error_details}")
            return JsonResponse({'response': f'ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {str(e)}'})
    
    return JsonResponse({'response': 'POSTãƒªã‚¯ã‚¨ã‚¹ãƒˆã®ã¿å¯¾å¿œã—ã¦ã„ã¾ã™ã€‚'})